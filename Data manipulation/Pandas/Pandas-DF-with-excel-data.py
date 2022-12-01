# import libraries
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt
#set the acive excel worksheets
wb13 = load_workbook("13.xlsx").active
wb16 = load_workbook("16.xlsx").active
# prepare the datasets
lst13 = []
lst16 = []
d = {}
# iterate through two files and extract the cell values
for col13, col16 in zip(wb13.iter_cols(min_row=4, min_col=2, max_row=7),
                        wb16.iter_cols(min_row=4, min_col=2, max_row=7)):

    for cell13, cell16 in zip(col13, col16):
        if cell13.value == None or cell16.value == None:
            continue
        # append the values to lists
        lst13.append(cell13.value)
        lst16.append(cell16.value)

# iterate through lists and calculate the difference in the numbers
for i, j in zip(range(0, len(lst13), 5), range(0, len(lst16), 5)):
    bxp = lst13[i]
    if bxp.startswith("ОБЩО") == False:
        # variables are made for readability
        exit13 = lst13[i + 2]
        enter13 = lst13[i + 4]
        exit16 = lst16[i + 2]
        enter16 = lst16[i + 4]
        exit_diff = exit16 - exit13
        enter_diff = enter16 - enter13
        
        # put all the info in dict if greater than 3 digit number
        if exit_diff > 100 and enter_diff > 100:
            d[bxp[4:]] = [exit_diff,enter_diff]

    else:
        # this calculates the "Total" cell values
        # again, variables are made for readability
        total_exit13 = lst13[i + 1]
        total_enter13 = lst13[i + 3]
        total_exit16 = lst16[i + 1]
        total_enter16 = lst16[i + 3]
        
        exit_diff = total_exit16  - total_exit13
        enter_diff = total_enter16 - total_enter13
        d[bxp[:4]] = [exit_diff, enter_diff]

# make the dataframe
df = pd.DataFrame.from_dict(d, orient='index', columns=['Exiting', 'Entering'])
#reset the index by default
df.reset_index(inplace=True)
# change the name of the columns, so 'BXP' is added
df.columns = ['BXP', 'Exiting', 'Entering']

# plot the data from DF with grid
ax = df.plot(x='BXP', y=['Exiting', 'Entering'], kind='bar', grid=True)
ax.bar_label(ax.containers[0], label_type='edge', rotation=90, padding=5)
ax.bar_label(ax.containers[1], label_type='edge', rotation=90, padding=5)
# sets the margins of subplots (figure is visible at all corners)
plt.subplots_adjust(left=0.05,
                    bottom=0.374,
                    right=0.98,
                    top=0.95,
                    wspace=0.4,
                    hspace=0.4)

# get current figure
figure = plt.gcf() 
# set the figure to be expanded, so it can be visible
figure.set_size_inches(13, 8)
# saves the figure as picture
# plt.savefig('foo.jpg')
# or show it
plt.show()
