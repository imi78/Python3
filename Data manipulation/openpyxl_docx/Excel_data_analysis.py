"""This program calculates the refugee's numbers from two Excel workbooks.
The first file must be named --> 13.xlxs with the figures form 24.02 to 31.12.2022.
The second file must be named --> 16.xlxs with the figures from 01.01.2023 onward.
The files have to be identical in terms of number of rows (from 1 to 7).
the 4th row must be the BXPs, 
5th row - "citizenship",
6th row - represents the exiting and entering PAX through the BXP column,
7th row - "Total"
in the 7th row ("Total") the new values will be written.
Any charts that you want to make should take the figures from the 7th row or the difference between 6th and 7th row.
The columns may vary as it reads only the identical text written in the 4th row starting from "B" column"""


# import libraries
from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt

#set the acive excel worksheets
wb13 = load_workbook("13.xlsx", data_only=True)
wb16 = load_workbook("16.xlsx", data_only=True)

# set the active worksheet
ws13 = wb13.active
ws16 = wb16.active
# initialize dictionary
d = {}

# function for sum up the values from both files and write them in wb13 and in the dictionary
def sum_values(bxp):
    # get the column letter and change it to int ("A" > 1)
    colnum13 = col13[0].column
    colnum16 = col16[0].column
    val13_exit, val16_exit = ws13.cell(row=6, column=colnum13), ws16.cell(row=6, column=colnum16)
    val13_enter, val16_enter = ws13.cell(row=6, column=colnum13+1), ws16.cell(row=6, column=colnum16+1)
    
    # sum the values
    sum_exit = val13_exit.value + val16_exit.value
    sum_enter = val13_enter.value + val16_enter.value
              
    # get the cells coordinates (e.g. "B7", "C7")
    cell_to_write_exit = ws13.cell(row=7, column=colnum13).coordinate
    cell_to_write_enter = ws13.cell(row=7, column=colnum13+1).coordinate

    # write the values to row No. 7
    # ws13["C7"] = value
    ws13[cell_to_write_exit] = sum_exit
    ws13[cell_to_write_enter] = sum_enter
    
    # put the difference from 7th and 6th row in a dictionary with:
    # val13_exit.value - row 6 value
    # sum_exit - total from row 7, 
    # and calculate the difference

    exit_diff = sum_exit-val13_exit.value
    enter_diff = sum_enter-val13_enter.value

    # only numbers above 1000...
    if exit_diff > 1000 and enter_diff > 1000:
        d[bxp] = [exit_diff, enter_diff]


for col13 in ws13.iter_cols(min_row=4, min_col=2, max_row=4):
    for col16 in ws16.iter_cols(min_row=4, min_col=2, max_row=4):
        
        # skip the merged cells
        if type(col13[0]).__name__ == 'MergedCell':
            continue
             
        # checks for empty cells
        if col13[0].value == None or col16[0].value == None:
            continue

        # check if the values are the same
        if col13[0].value == col16[0].value:
            # calculate the "Total" cells
            if col13[0].value.startswith("ОБЩО"):
                bxp = col13[0].value
                bxp = bxp[:4]
		        # function call
                sum_values(bxp)
                break
            else:
                bxp = col13[0].value
                sum_values(bxp)
                
    else:  # only execute when it's no break in the inner loop
        continue
    break # breaks the outer loop
    
# save the workbook
wb13.save('test.xlsx')

"""if you want a chart to be saved, uncomment next lines of code"""

# # make the dataframe
# df = pd.DataFrame.from_dict(d, orient='index', columns=['Exiting', 'Entering'])
# #reset the index by default
# df.reset_index(inplace=True)
# # change the name of the columns, so 'BXP' is added
# df.columns = ['BXP', 'Exiting', 'Entering']

# # plot the data from DF with grid
# ax = df.plot(x='BXP', y=['Exiting', 'Entering'], kind='bar', grid=True)

# # sets the bar labels to be rotated for visibility
# ax.bar_label(ax.containers[0], label_type='edge', rotation=90, padding=5)
# ax.bar_label(ax.containers[1], label_type='edge', rotation=90, padding=5)

# # sets the margins of subplots (figure is visible at all corners)
# plt.subplots_adjust(left=0.05,
#                     bottom=0.374,
#                     right=0.98,
#                     top=0.95,
#                     wspace=0.4,
#                     hspace=0.4)

# # get current figure

# figure = plt.gcf() 
# # set the figure to be expanded, so it can be visible

# figure.set_size_inches(13, 8)

# # saves the figure as picture
# plt.savefig('foo.jpg')

# # or show it
# # plt.show()
